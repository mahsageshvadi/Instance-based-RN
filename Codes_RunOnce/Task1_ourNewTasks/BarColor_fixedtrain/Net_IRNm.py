import numpy as np
import keras
import keras.backend as K
from keras.models import Sequential, Model
from keras.layers import Dense, Dropout, Flatten,Input
from keras.layers import Conv2D, MaxPooling2D
import tensorflow as tf
from keras.optimizers import SGD, Adam
import matplotlib.pyplot as plt
import cv2
import os
from  openpyxl import Workbook
import sklearn
from sklearn.metrics import mean_squared_error
from Configure import Config, GetProcessBar,ReadLinesInFile, ClearDir, MakeDir, RemoveDir
import time, argparse
from utils.non_local import non_local_block


""" some important configuration """
train_num = 60000             # image number.
val_num   = 20000
test_num  = 20000

m_epoch = 100                # epoch
m_batchSize = 32            # batch_size
m_print_loss_step = 15      # print once after how many iterations.



""" processing command line """
parser = argparse.ArgumentParser()
parser.add_argument("--lr", default=0.0001, type = float)  # learning rate
parser.add_argument("--gpu", default='0')                  # gpu id
parser.add_argument("--savedir", default= 'IRNm')           # saving path.
parser.add_argument("--backup", default=False, type=bool)   # whether to save weights after each epoch.
                                                           # (If True, it will cost lots of memories)
a = parser.parse_args()

m_optimizer = Adam(a.lr)
#os.environ["CUDA_VISIBLE_DEVICES"] = a.gpu

config = Config()

# create save folder.
MakeDir("./results/")
dir_results   = os.path.abspath(".") + "/results/{}/".format(a.savedir)    # backup path. 'results/IRN_m'




def GetInformation(flag):
    if flag == 'train':
        return train_num, config.dir_subCharts_train, config.path_groundTruth_train
    if flag == 'val':
        return val_num, config.dir_subCharts_val, config.path_groundTruth_val
    if flag == 'test':
        return test_num, config.dir_subCharts_test, config.path_groundTruth_test

def LoadSeparateChartDataSet(flag = 'train'):  # must be 'train' or 'val' or 'test'

    __max_load_num, __dirSubPath, __path_groundTruth = GetInformation(flag)

    # load the segmented sub-chart images.
    images = np.ones((config.max_obj_num, __max_load_num , config.image_height, config.image_width, 3), dtype='float32')

    old_curpath = os.path.abspath(".")
    os.chdir(__dirSubPath)  # change current path to load images faster.
    count = 0
    for imgID in range(__max_load_num):
        for barID in range(config.max_obj_num):
            imagePath = config.subChartName.format(imgID, barID)

            if os.path.exists(imagePath):
                images[barID][imgID] = cv2.imread(imagePath) / 255.0

        if count % 10000 == 0:
            print("Loaded: {}/{}".format(count, __max_load_num))
        count += 1

    # read the labels.
    labels = []
    lines = ReadLinesInFile(__path_groundTruth)  # all lines in ground-truth file.
    for line in lines:
        elements = line.split()
        if len(elements) == 0:
            continue
        if len(labels) >= __max_load_num:
            break
        labels.append([float(p) for p in elements])

    os.chdir(old_curpath)
    # print("cur path", os.path.abspath(os.curdir))

    labels = np.array(labels,dtype='float32')

    print("---------------------------")
    print("[Process] x: ", images.shape)
    print("[Process] y: ", labels.shape, config.max_obj_num)

    return images, labels



# Level1 module is to extract the individual features from one instance.
# Level1 has two NON-LOCAL block.
def Level1_Module():
    input = Input(shape=(config.image_height, config.image_width, 3))
    x = Conv2D(32, (3, 3), activation='relu', padding='same')(input)
    x = Conv2D(32, (3, 3), activation='relu', padding='same')(x)
    x = MaxPooling2D(pool_size=(2, 2), strides=(2, 2))(x)
    x = Conv2D(64, (3, 3), activation='relu', padding='same')(x)
    x = Conv2D(64, (3, 3), activation='relu', padding='same')(x)
    x = MaxPooling2D(pool_size=(2, 2), strides=(2, 2))(x)
    x = non_local_block(x)   # non local block
    x = Conv2D(64, (3, 3), activation='relu', padding='same')(x)
    x = Conv2D(64, (3, 3), activation='relu', padding='same')(x)
    x = Conv2D(64, (3, 3), activation='relu', padding='same')(x)
    x = MaxPooling2D(pool_size=(2, 2), strides=(2, 2))(x)
    x = non_local_block(x)   # non local block
    return Model(inputs=input, outputs=x)

# Level2 module is to compute the ratio of a pair.
# Level2 has one NON-LOCAL block.
def Level2_Module(w,h,c):
    print("Level2:", w,h,c)

    inputA = Input(shape=(w, h, c))
    inputB = Input(shape=(w, h, c))

    combined = keras.layers.concatenate([inputA, inputB])   # concatenate them.
    z = Conv2D(64, (3, 3), activation='relu',padding='same')(combined)
    z = Conv2D(64, (3, 3), activation='relu', padding='same')(z)
    z = non_local_block(z)   # non local block
    #
    z = Flatten()(z)
    z = Dense(256, activation="relu")(z)
    z = Dropout(0.5)(z)
    z = Dense(1, activation="linear")(z)  # output the ratio of this pair.

    return Model(inputs=[inputA, inputB], outputs=z)



# IRN_m is final network to estimate the ratio vectors from multiple input instances.
def Build_IRN_m_Network():
    # input layers.
    input_layers = []
    # the first 'obj_num' inputs are corresponding to the input sub-charts.
    for i in range(config.max_obj_num):
        input = Input(shape=(config.image_height, config.image_width, 3), name="input_{}".format(i))
        input_layers.append(input)

    # The last input layer is used for representing R1=(o1/o1)=1.0 which is just a constant.
    # Here, I would use an extra input layer which is 1-dim and always equal to 1.0 rather than directly using a contant.
    # It makes same effect and can avoid some strange compile errors. (I only use TensorFlow before, not way familiar to Keras.)
    R1_one_input = Input(shape=(1,),name="input_constant_scalar1",dtype='float32')   # always equal to 1.0.
    input_layers.append(R1_one_input)

    # First extract individual features.
    individual_features = []
    level1 = Level1_Module()  # build a level1 module
    for i in range(config.max_obj_num):
        x = level1(input_layers[i])
        individual_features.append(x)

    # Use a Level2 module to predict pairwise ratios.
    level2 = Level2_Module(w=int(individual_features[0].shape[1]),
                           h=int(individual_features[0].shape[2]),
                           c=int(individual_features[0].shape[3]))

    ratio_p_layers = [R1_one_input]   # pairwise ratio vector. put in' R1=(o1/o1)=1.0 '.
    for i in range(config.max_obj_num-1): # compute the ratio of each neighbor pair.
        x = level2(inputs = [individual_features[i], individual_features[i+1]])
        ratio_p_layers.append(x)

    print("ratio_p_layers", len(ratio_p_layers), ratio_p_layers[-1].shape)

    # Compute the ratios relative to the first object by using MULTIPLY() operation.
    ratio_layers = [R1_one_input]  # put in R1=1.0.
    i = 1
    while i<len(ratio_p_layers):
        x = keras.layers.Multiply()(ratio_p_layers[:i+1])   # R1*R2*...Ri
        i+=1
        ratio_layers.append(x)

    # divide the maxinum of 'ratio_layers' to get the final results.
    max = keras.layers.maximum(ratio_layers)
    z = keras.layers.concatenate(ratio_layers)
    z = keras.layers.Lambda(lambda x: x[0]/x[1])([z, max])

    print("output layer: ", z.shape)

    return Model(inputs=input_layers, outputs=z)

# save the 'predicted results' and 'ground truth' into the file.
def SavePredictedResult(x, y, flag = 'train'):
    dim = y.shape[1]
    print(y.shape)
    input_test = [x[i] for i in range(config.max_obj_num)]
    input_test.append(np.ones(x[0].shape[0]))
    predict_Y = model.predict(x=input_test, batch_size=1)
    predictFile = open(dir_results + flag + "_predicted_results.txt",'w')
    for i in range(y.shape[0]):
        for t in range(dim):  # save the ground_truth
            predictFile.write(str(y[i,t]) + '\t')
        predictFile.write('\n')
        for t in range(dim):  # save the predicted results.
            predictFile.write(str(predict_Y[i, t]) + '\t')
        predictFile.write('\n')
    predictFile.close()

    MLAE = np.log2(sklearn.metrics.mean_absolute_error( predict_Y * 100, y * 100) + .125)

    return MLAE

if __name__ == '__main__':
    ClearDir(dir_results)
    ClearDir(dir_results+"backup")

    x_train, y_train = LoadSeparateChartDataSet(flag='train')
    x_val, y_val = LoadSeparateChartDataSet(flag='val')
    x_test, y_test = LoadSeparateChartDataSet(flag='test')
    x_train -= .5
    x_val -= .5
    x_test -= .5

    # format the inputs as [img1, img2, img3,....., 1.0]
    # Note that we add R0=1.0 in the last.
    x_train = [x_train[i] for i in range(config.max_obj_num)]
    x_val   = [x_val[i] for i in range(config.max_obj_num)]
    x_test  = [x_test[i] for i in range(config.max_obj_num)]
    x_train.append(np.ones(train_num))   # Train_num, R0 = 1.0
    x_val.append(np.ones(val_num))       # Val_num, R0 = 1.0
    x_test.append(np.ones(test_num))     # Test_num, R0 = 1.0



    print("----------Build Network----------------")
    model = Build_IRN_m_Network()
    model.compile(loss='mse', optimizer=m_optimizer)

    print("----------Training-------------------")
    history_batch = []
    history_iter = []
    batch_amount = train_num // m_batchSize
    rest_size = train_num - (batch_amount*m_batchSize)

    # best model according to the training loss. (Since this network can't deal with this generalization task)
    best_train_loss = 99999.99999
    val_loss_using_Train = 99999.99999    # corresponding val loss using the best model on training set.
    best_model_name_onTrain = "xxxxx"

    # best model according to val loss.
    best_val_loss = 99999.999999
    best_model_name_onVal = "xxxxx"


    for iter in range(m_epoch):

        # shuffle the training set !!!!!!!!!!!!!!
        index = [i for i in range(train_num)]
        np.random.shuffle(index)

        for bid in range(batch_amount):

            # using the shuffle index...
            x_batch = [ x_train[i][index[bid*m_batchSize : (bid + 1)*m_batchSize]] for i in range(config.max_obj_num+1)]
            y_batch = y_train[index[bid*m_batchSize : (bid + 1)*m_batchSize]]

            model.train_on_batch(x= x_batch, y=y_batch)  # training on batch

            if bid % m_print_loss_step == 0:
                logs = model.evaluate(x=x_batch,y=y_batch,verbose=0, batch_size=m_batchSize)
                print("iter({}/{}) Batch({}/{}) {} : mse_loss={}".format(iter, m_epoch, bid, batch_amount,
                                                                     GetProcessBar(bid, batch_amount), logs))
                history_batch.append([iter, bid, logs])

        # training on the rest data.
        if rest_size > 0:
            model.train_on_batch(x=[ x_train[i][index[batch_amount*m_batchSize : ]] for i in range(config.max_obj_num+1)],
                                 y=y_train[index[batch_amount*m_batchSize : ]])

        # one epoch is done. Do some information collections.
        epoch_loss_train = model.evaluate(x=x_train, y=y_train, verbose=0, batch_size=m_batchSize)
        epoch_loss_val = model.evaluate(x=x_val, y=y_val, verbose=0, batch_size=m_batchSize)
        history_iter.append([iter, epoch_loss_train, epoch_loss_val])
        print("----- epoch({}/{}) train_loss={}, val_loss={}".format(iter, m_epoch, epoch_loss_train, epoch_loss_val))

        # if a.backup == True:   # whether to save the weight after each epoch
        #     model.save_weights(dir_results+"backup"+"/model_{}_{}.h5".format(iter, epoch_loss_val))

        if epoch_loss_train < best_train_loss:
            RemoveDir(best_model_name_onTrain)
            best_model_name_onTrain = dir_results + "model_irnm_onTrain_{}.h5".format(epoch_loss_val)
            model.save_weights(best_model_name_onTrain)
            val_loss_using_Train = epoch_loss_val
            best_train_loss = epoch_loss_train

        if epoch_loss_val < best_val_loss:  # save the best model on Validation set.
            best_val_loss = epoch_loss_val
            RemoveDir(best_model_name_onVal)
            best_model_name_onVal = dir_results + "model_irnm_onVal_{}.h5".format(epoch_loss_val)
            model.save_weights(best_model_name_onVal)



    # test on the testing set.
    model.load_weights(best_model_name_onTrain)   # using the best model
    test_loss_usingTrain = model.evaluate(x_test, y_test, verbose=0, batch_size=m_batchSize)


    # Save the predicted results and return the MLAE.
    MLAE_train = SavePredictedResult(x_train,y_train,'train_usingTrain')
    MLAE_val = SavePredictedResult(x_val, y_val, 'val_usingTrain')
    MLAE_test = SavePredictedResult(x_test,y_test,'test_usingTrain')

    model.load_weights(best_model_name_onVal)   # using the best model.
    train_loss_onVal = model.evaluate(x_train, y_train, verbose=0, batch_size=m_batchSize)
    test_loss_onVal = model.evaluate(x_test, y_test, verbose=0, batch_size=m_batchSize)
    MLAE_train_onVal = SavePredictedResult(x_train,y_train,'train_usingVal')
    MLAE_val_onVal = SavePredictedResult(x_val, y_val, 'val_usingVal')
    MLAE_test_onVal = SavePredictedResult(x_test, y_test, 'test_usingVal')


    # save the training information.
    wb = Workbook()
    ws1 = wb.active                         # MSE/MLAE
    ws1.title = "MLAE_MSE"
    ws2 = wb.create_sheet("EPOCH loss")      # iteration loss
    ws3 = wb.create_sheet("BATCH loss")     # batch loss

    ws2.append(["Epoch ID", "Train MSE Loss", "Val MSE Loss"])
    ws3.append(["Epoch ID", "Batch ID", "MSE Loss"])

    for i in range(len(history_iter)):
        ws2.append(history_iter[i])
    for i in range(len(history_batch)):
        ws3.append(history_batch[i])

    ws1.append(["----------", "Using best model on train_set"])
    ws1.append(["Best Train loss", best_train_loss])
    ws1.append(["Val loss usingTrain", val_loss_using_Train])
    ws1.append(["Test loss usingTrain", test_loss_usingTrain])
    ws1.append(["Train MLAE", MLAE_train])
    ws1.append(["val MLAE", MLAE_val])
    ws1.append(["Test MLAE", MLAE_test])
    ws1.append(["----------", "Using best model on val_set    !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!"])
    ws1.append(["Train loss usingVal", train_loss_onVal])
    ws1.append(["Best Val loss", best_val_loss])
    ws1.append(["Test loss usingVal", test_loss_onVal])
    ws1.append(["Train MLAE using Val", MLAE_train_onVal])
    ws1.append(["Val MLAE using Val", MLAE_val_onVal])
    ws1.append(["Test MLAE using Val", MLAE_test_onVal])

    wb.save(dir_results + "train_info.xlsx")

    print("-----Using the best model according to training loss-------")
    print("Training MSE:", best_train_loss)
    print("Validat. MSE", val_loss_using_Train)
    print("Testing MSE:", test_loss_usingTrain)
    print("Training MLAE:", MLAE_train)
    print("Validat. MLAE", MLAE_val)
    print("Testing MLAE:", MLAE_test)

    print("-----Using the best model according to Validation loss-------")
    print("Training MSE:", train_loss_onVal)
    print("Validat. MSE:", best_val_loss)
    print("Testing MSE:", test_loss_onVal)
    print("Training MLAE:", MLAE_train_onVal)
    print("Validat. MLAE", MLAE_val_onVal)
    print("Testing MLAE:", MLAE_test_onVal)
